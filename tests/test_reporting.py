from __future__ import annotations

import json
from datetime import UTC, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from aruba2930f_backup.reporting import (
    ReportSummary,
    SanitizedJsonlLogger,
    sanitize_log_event,
    write_result_workbook,
)


def test_result_workbook_has_expected_sheets_fields_and_formula_protection(tmp_path: Path) -> None:
    started = datetime(2026, 8, 20, 1, 2, 3, tzinfo=UTC)
    result = {
        "target": {"ip": "192.0.2.10"},
        "hostname": '=HYPERLINK("https://invalid.example")',
        "model": "Aruba 2930F",
        "sku": "JL253A",
        "status": "success",
        "host_key_attempts": 1,
        "attempts": 1,
        "started_at": started,
        "finished_at": started + timedelta(seconds=2),
        "config_path": tmp_path / "edge.txt",
        "config_sha256": "a" * 64,
        "diagnostic_code": "A3F1-010EPMRC-3",
        "error_code": "",
        "error_message": "+untrusted formula text",
    }

    report_path = write_result_workbook(
        tmp_path,
        [result],
        summary=ReportSummary(
            started_at=started,
            finished_at=started + timedelta(seconds=2),
        ),
    )

    workbook = load_workbook(report_path, data_only=False)
    assert workbook.sheetnames == ["Summary", "Devices"]
    devices = workbook["Devices"]
    assert [cell.value for cell in devices[1]] == [
        "IP",
        "Hostname",
        "Model/SKU",
        "Status",
        "Host Key Attempts",
        "Backup Attempts",
        "Total Connection Attempts",
        "Started At",
        "Finished At",
        "Duration Seconds",
        "Config File",
        "SHA-256",
        "Diagnostic Code",
        "Error Code",
        "Error Message",
    ]
    assert devices["A2"].value == "192.0.2.10"
    assert devices["B2"].value.startswith("'=")
    assert devices["C2"].value == "Aruba 2930F / JL253A"
    assert devices["M2"].value == "A3F1-010EPMRC-3"
    assert devices["O2"].value.startswith("'+")
    assert devices["D2"].value == "success"
    assert devices["E2"].value == 1
    assert devices["F2"].value == 1
    assert devices["G2"].value == 2
    assert workbook["Summary"]["B5"].value == 1
    assert not (tmp_path / "result.xlsx.part").exists()
    workbook.close()


def test_result_workbook_preserves_family_only_and_vsf_model_display(tmp_path: Path) -> None:
    report_path = write_result_workbook(
        tmp_path,
        [
            {
                "target": {"ip": "192.0.2.10"},
                "model": "Aruba 2930F",
                "sku": None,
                "status": "success",
            },
            {
                "target": {"ip": "192.0.2.11"},
                "model": "Aruba 2930F VSF",
                "sku": "JL253A, JL255A",
                "status": "success",
            },
        ],
    )

    workbook = load_workbook(report_path, data_only=True)
    devices = workbook["Devices"]
    assert devices["C2"].value == "Aruba 2930F"
    assert devices["C3"].value == "Aruba 2930F VSF / JL253A, JL255A"
    workbook.close()


def test_jsonl_logger_allowlists_fields_and_redacts_sensitive_values(tmp_path: Path) -> None:
    path = tmp_path / "operation.jsonl"
    logger = SanitizedJsonlLogger(
        path,
        sensitive_values=("operator", "secret-password", "edge-switch"),
    )

    returned = logger.log(
        stage="connecting",
        status="failed",
        error_code="TCP_TIMEOUT",
        diagnostic_code="A3F1-010EPMRC-3",
        message="operator at 192.0.2.10 edge-switch secret-password\nnext line",
        ip="192.0.2.10",
        password="secret-password",
        raw_output="running configuration",
    )

    record = json.loads(path.read_text(encoding="utf-8"))
    assert record == returned
    encoded = json.dumps(record)
    assert "192.0.2.10" not in encoded
    assert "secret-password" not in encoded
    assert "operator" not in encoded
    assert "edge-switch" not in encoded
    assert "raw_output" not in record
    assert "password" not in record
    assert record["diagnostic_code"] == "A3F1-010EPMRC-3"
    assert record["message"].count("<redacted") >= 4


def test_sanitize_log_event_does_not_serialize_arbitrary_object_fields() -> None:
    event = {
        "stage": "reading_config",
        "attempt": 1,
        "config_text": "password manager plaintext",
        "target": {"ip": "198.51.100.2"},
    }

    record = sanitize_log_event(event)

    assert record["stage"] == "reading_config"
    assert record["attempt"] == 1
    assert "config_text" not in record
    assert "target" not in record


def test_retry_exhausted_is_counted_separately(tmp_path: Path) -> None:
    report_path = write_result_workbook(
        tmp_path,
        [
            {
                "target": {"ip": "192.0.2.10"},
                "status": "retry_exhausted",
                "host_key_attempts": 4,
                "attempts": 0,
            },
            {"target": {"ip": "192.0.2.11"}, "status": "failed"},
        ],
    )

    workbook = load_workbook(report_path, data_only=True)
    summary = {
        workbook["Summary"].cell(row=row, column=1).value: workbook["Summary"]
        .cell(row=row, column=2)
        .value
        for row in range(2, workbook["Summary"].max_row + 1)
    }
    assert summary["Total Devices"] == 2
    assert summary["Successful"] == 0
    assert summary["Failed"] == 1
    assert summary["Retry Exhausted"] == 1
    assert summary["Cancelled"] == 0
    workbook.close()


def test_jsonl_logger_preserves_retry_round_and_delay_but_not_target(tmp_path: Path) -> None:
    path = tmp_path / "operation.jsonl"
    logger = SanitizedJsonlLogger(path, sensitive_values=("192.0.2.10",))

    logger.log(
        phase="host_key",
        stage="retry_wait",
        round=2,
        attempt=1,
        delay_seconds=5.0,
        error_code="TCP_TIMEOUT",
        message="waiting for 192.0.2.10",
        target={"ip": "192.0.2.10"},
    )

    record = json.loads(path.read_text(encoding="utf-8"))
    assert record["stage"] == "retry_wait"
    assert record["phase"] == "host_key"
    assert record["round"] == 2
    assert record["attempt"] == 1
    assert record["delay_seconds"] == 5.0
    assert record["error_code"] == "TCP_TIMEOUT"
    assert "target" not in record
    assert "192.0.2.10" not in json.dumps(record)

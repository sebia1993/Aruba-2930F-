"""Excel and sanitized operational reporting."""

from __future__ import annotations

import json
import os
import re
import threading
from collections.abc import Iterable, Mapping
from dataclasses import asdict, dataclass, is_dataclass
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DEVICE_HEADERS = (
    "IP",
    "Hostname",
    "Model/SKU",
    "Status",
    "Host Key Attempts",
    "Backup Attempts",
    "Total Connection Attempts",
    "Started At",
    "Finished At",
    "Duration Seconds",
    "Config File",
    "SHA-256",
    "Diagnostic Code",
    "Error Code",
    "Error Message",
)

_FORMULA_PREFIXES = ("=", "+", "-", "@")
_IPV4_TEXT = re.compile(r"(?<![\w.])(?:\d{1,3}\.){3}\d{1,3}(?![\w.])")
_CONTROL_TEXT = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
_LOG_ALLOWED_FIELDS = {
    "timestamp",
    "level",
    "phase",
    "stage",
    "status",
    "error_code",
    "diagnostic_code",
    "round",
    "attempt",
    "delay_seconds",
    "elapsed_ms",
    "duration_ms",
    "retryable",
    "message",
    "count",
    "completed",
    "total",
}


@dataclass(frozen=True, slots=True)
class ReportSummary:
    """High-level report metadata used by ``write_result_workbook``."""

    started_at: datetime | None = None
    finished_at: datetime | None = None
    cancelled: bool = False


def _object_mapping(value: object) -> Mapping[str, Any]:
    if isinstance(value, Mapping):
        return value
    if is_dataclass(value) and not isinstance(value, type):
        converted = asdict(value)
        if isinstance(converted, Mapping):
            return converted
    try:
        return vars(value)
    except TypeError:
        return {}


def _read_path(value: object, path: str) -> Any:
    current: Any = value
    for part in path.split("."):
        if isinstance(current, Mapping):
            if part not in current:
                return None
            current = current[part]
        else:
            current = getattr(current, part, None)
        if current is None:
            return None
    return current


def _first(value: object, *paths: str, default: Any = "") -> Any:
    for path in paths:
        candidate = _read_path(value, path)
        if candidate is not None:
            return candidate
    return default


def _plain(value: Any) -> Any:
    if isinstance(value, Enum):
        return value.value
    if isinstance(value, Path):
        return str(value)
    return value


def neutralize_excel_text(value: Any) -> Any:
    """Prevent untrusted device text from becoming an Excel formula."""

    value = _plain(value)
    if isinstance(value, datetime) and value.tzinfo is not None:
        # Excel stores wall-clock values and rejects timezone-aware datetimes.
        value = value.astimezone().replace(tzinfo=None)
    if not isinstance(value, str):
        return value
    value = _CONTROL_TEXT.sub("", value)
    if value.lstrip().startswith(_FORMULA_PREFIXES):
        return f"'{value}"
    return value


def _duration_seconds(result: object) -> float | str:
    explicit = _first(result, "duration_seconds", "elapsed_seconds", default=None)
    if explicit is not None:
        try:
            return round(float(explicit), 3)
        except TypeError, ValueError:
            return str(neutralize_excel_text(explicit))
    started = _first(result, "started_at", default=None)
    finished = _first(result, "finished_at", default=None)
    if isinstance(started, datetime) and isinstance(finished, datetime):
        return round((finished - started).total_seconds(), 3)
    return ""


def _model_sku(result: object) -> str:
    combined = _first(result, "model_sku", "model_and_sku", default="")
    if combined:
        return str(combined)
    model = str(_first(result, "model", default="") or "")
    sku = str(_first(result, "sku", "product_number", default="") or "")
    return " / ".join(part for part in (model, sku) if part)


def _device_row(result: object) -> tuple[Any, ...]:
    host_key_attempts = _first(result, "host_key_attempts", default=0)
    backup_attempts = _first(result, "attempts", "attempt_count", default=0)
    try:
        total_attempts: int | str = int(host_key_attempts) + int(backup_attempts)
    except TypeError, ValueError:
        total_attempts = ""
    return (
        _first(result, "ip_address", "ip", "target.ip_address", "target.ip"),
        _first(result, "hostname", "device_hostname"),
        _model_sku(result),
        _first(result, "status"),
        host_key_attempts,
        backup_attempts,
        total_attempts,
        _first(result, "started_at"),
        _first(result, "finished_at"),
        _duration_seconds(result),
        _first(result, "config_path", "file_path", "stored.path"),
        _first(result, "sha256", "config_sha256", "stored.sha256"),
        _first(result, "diagnostic_code"),
        _first(result, "error_code", "error.code"),
        _first(result, "error_message", "message", "error.message"),
    )


def _status_text(result: object) -> str:
    value = _plain(_first(result, "status", default=""))
    return str(value).upper()


def _style_sheet(sheet: Any, *, widths: tuple[int, ...]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def write_result_workbook(
    destination: str | os.PathLike[str],
    results: Iterable[object],
    *,
    summary: ReportSummary | None = None,
) -> Path:
    """Write ``result.xlsx`` atomically and return its path."""

    result_list = list(results)
    report_summary = summary or ReportSummary()
    destination_path = Path(destination)
    report_path = (
        destination_path
        if destination_path.suffix.lower() == ".xlsx"
        else destination_path / "result.xlsx"
    )
    report_path.parent.mkdir(parents=True, exist_ok=True)

    statuses = [_status_text(result) for result in result_list]
    successes = sum(status in {"SUCCESS", "SUCCEEDED", "COMPLETED"} for status in statuses)
    retry_exhausted = sum(status == "RETRY_EXHAUSTED" for status in statuses)
    cancellations = sum(status == "CANCELLED" for status in statuses)
    failures = len(result_list) - successes - retry_exhausted - cancellations
    duration: float | str = ""
    if report_summary.started_at and report_summary.finished_at:
        duration = round(
            (report_summary.finished_at - report_summary.started_at).total_seconds(),
            3,
        )

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(("Field", "Value"))
    summary_values = (
        ("Run Start", report_summary.started_at),
        ("Run Finish", report_summary.finished_at),
        ("Duration Seconds", duration),
        ("Total Devices", len(result_list)),
        ("Successful", successes),
        ("Failed", failures),
        ("Retry Exhausted", retry_exhausted),
        ("Cancelled", cancellations),
        ("Run Cancelled", report_summary.cancelled),
        ("Result Directory", str(report_path.parent)),
    )
    for field, value in summary_values:
        summary_sheet.append((field, neutralize_excel_text(value)))
    _style_sheet(summary_sheet, widths=(24, 64))

    devices_sheet = workbook.create_sheet("Devices")
    devices_sheet.append(DEVICE_HEADERS)
    for result in result_list:
        devices_sheet.append(tuple(neutralize_excel_text(value) for value in _device_row(result)))
    _style_sheet(
        devices_sheet,
        widths=(16, 24, 30, 18, 18, 18, 24, 21, 21, 18, 48, 66, 24, 26, 54),
    )
    for row in devices_sheet.iter_rows(min_row=2):
        row[7].number_format = "yyyy-mm-dd hh:mm:ss"
        row[8].number_format = "yyyy-mm-dd hh:mm:ss"
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    partial_path = report_path.with_name(f"{report_path.name}.part")
    try:
        workbook.save(partial_path)
        os.replace(partial_path, report_path)
    except BaseException:
        partial_path.unlink(missing_ok=True)
        raise
    finally:
        workbook.close()
    return report_path


def _sanitize_message(value: Any, sensitive_values: tuple[str, ...]) -> str:
    message = str(_plain(value) if value is not None else "")
    message = _CONTROL_TEXT.sub("", message.replace("\r", " ").replace("\n", " "))
    message = _IPV4_TEXT.sub("<redacted-ip>", message)
    for sensitive in sensitive_values:
        if sensitive:
            message = message.replace(sensitive, "<redacted>")
    return message[:500]


def sanitize_log_event(
    event: object | None = None,
    *,
    sensitive_values: Iterable[str] = (),
    **fields: Any,
) -> dict[str, Any]:
    """Return an allowlisted, single-line operational log record."""

    source = dict(_object_mapping(event)) if event is not None else {}
    source.update(fields)
    secrets = tuple(str(value) for value in sensitive_values if value is not None)
    sanitized: dict[str, Any] = {
        "timestamp": datetime.now().astimezone().isoformat(timespec="seconds")
    }
    for key in _LOG_ALLOWED_FIELDS:
        if key == "timestamp" or key not in source:
            continue
        value = _plain(source[key])
        if key == "message":
            sanitized[key] = _sanitize_message(value, secrets)
        elif isinstance(value, (str, int, float, bool)) or value is None:
            sanitized[key] = value
        else:
            sanitized[key] = str(value)
    return sanitized


class SanitizedJsonlLogger:
    """Thread-safe writer for non-sensitive stage/error diagnostics."""

    def __init__(
        self,
        path: str | os.PathLike[str],
        *,
        sensitive_values: Iterable[str] = (),
    ) -> None:
        self.path = Path(path)
        self._sensitive_values = tuple(str(value) for value in sensitive_values)
        self._lock = threading.Lock()

    def log(self, event: object | None = None, **fields: Any) -> dict[str, Any]:
        record = sanitize_log_event(
            event,
            sensitive_values=self._sensitive_values,
            **fields,
        )
        encoded = json.dumps(record, ensure_ascii=False, separators=(",", ":"))
        self.path.parent.mkdir(parents=True, exist_ok=True)
        with self._lock, self.path.open("a", encoding="utf-8", newline="\n") as stream:
            stream.write(encoded)
            stream.write("\n")
            stream.flush()
            os.fsync(stream.fileno())
        return record
